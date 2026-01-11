#!/usr/bin/env python3
"""
Script para visualizar a estrutura da planilha Excel gerada
Mostra uma prévia em texto da estrutura da tabela
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def visualizar_estrutura():
    wb = load_workbook('CHECK_LIST_EQUIPAMENTOS_SALA_01.xlsx')
    ws = wb.active
    
    print("=" * 100)
    print("VISUALIZAÇÃO DA ESTRUTURA DA PLANILHA")
    print("=" * 100)
    print()
    
    # Título e Mês/Ano
    print(f"TÍTULO: {ws['A1'].value}")
    print(f"CAMPO DATA: {ws['A2'].value}")
    print()
    print("-" * 100)
    
    # Cabeçalho da tabela
    print(f"{'MATERIAL':<30} | ", end="")
    for dia in range(1, 32):
        print(f"{dia:>2} ", end="")
        if dia < 31:
            print("  ", end="")
    print()
    
    # Subcabeçalho (M, T, N)
    print(f"{'':<30} | ", end="")
    for dia in range(1, 32):
        print("MTN", end="")
        if dia < 31:
            print("  ", end="")
    print()
    print("-" * 100)
    
    # Equipamentos
    equipamentos = [
        "ELETROCAUTÉRIO",
        "MESA CIRÚRGICA TOMADA",
        "CARRO ANEST./CAPNÓGRAFO",
        "MONITOR COMPLETO",
        "OX PERFURO CORTANTE",
        "OX E LÂMINAS LARINGO",
        "ESTOJO AD E INF",
        "MÓDULO P.A.I",
        "PAREDE GASES FUNCIONANDO",
        "CIRCUITO ANESTESIA /KTS",
        "02 SUPORTE SORO",
        "03 MESAS + 01 MAYO",
        "REANIMADOR AD E INF.",
        "02 LIXEIRAS",
        "HAMPER"
    ]
    
    linha = 5
    for equip in equipamentos:
        print(f"{equip:<30} | [Campos para preenchimento diário]")
        linha += 1
    
    print("-" * 100)
    print(f"{'VISTO DO PROFISSIONAL':<30} | [Campos para visto diário]")
    print("-" * 100)
    print()
    
    # Informações adicionais
    print("INFORMAÇÕES ADICIONAIS:")
    print(f"  - Total de linhas de equipamentos: 15")
    print(f"  - Total de dias: 31")
    print(f"  - Turnos por dia: 3 (Manhã, Tarde, Noite)")
    print(f"  - Total de células de checagem: 31 × 3 × 16 = 1,488 células")
    print(f"  - Orientação: PAISAGEM")
    print(f"  - Formato: Excel XLSX")
    print()
    print("=" * 100)
    print("A planilha está pronta para uso e impressão!")
    print("=" * 100)

if __name__ == "__main__":
    visualizar_estrutura()
