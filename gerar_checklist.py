#!/usr/bin/env python3
"""
Script para gerar planilha Excel de CHECK LIST DE EQUIPAMENTOS
para Centro Cirúrgico - Sala 01
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Constantes
TOTAL_DIAS = 31
TURNOS_POR_DIA = 3
TOTAL_COLUMNS = 2 + (TOTAL_DIAS * TURNOS_POR_DIA)  # 1 coluna MATERIAL + 1 offset + 93 colunas de dias

def criar_checklist():
    # Criar novo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist"
    
    # Lista de equipamentos
    equipamentos = [
        "ELETROCAUTÉRIO",
        "MESA CIRÚRGICA TOMADA",
        "CARRO ANEST./CAPNÓGRAFO",
        "MONITOR COMPLETO",
        "OX PERFURO CORTANTE",
        "OX E LÂMINAS LARINGO",
        "ESTOJO AD E INF",
        "MÓDULO P.A.I",
        "PAREDE GASES FUNCIONANDO",
        "CIRCUITO ANESTESIA /KTS",
        "02 SUPORTE SORO",
        "03 MESAS + 01 MAYO",
        "REANIMADOR AD E INF.",
        "02 LIXEIRAS",
        "HAMPER"
    ]
    
    # Configurar bordas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Linha 1: Título
    ws.merge_cells('A1:CF1')
    titulo_cell = ws['A1']
    titulo_cell.value = "CHECK LIST DE EQUIPAMENTOS – CENTRO CIRÚRGICO – SALA 01"
    titulo_cell.font = Font(bold=True, size=14)
    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Linha 2: Mês/Ano
    ws.merge_cells('A2:CF2')
    mes_ano_cell = ws['A2']
    mes_ano_cell.value = "MÊS/ANO: __________"
    mes_ano_cell.font = Font(bold=True, size=12)
    mes_ano_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Linha 3: Cabeçalho da tabela
    linha_cabecalho = 3
    
    # Célula MATERIAL
    ws.merge_cells(f'A{linha_cabecalho}:A{linha_cabecalho+1}')
    material_cell = ws[f'A{linha_cabecalho}']
    material_cell.value = "MATERIAL"
    material_cell.font = Font(bold=True, size=11)
    material_cell.alignment = Alignment(horizontal='center', vertical='center')
    material_cell.border = thin_border
    
    # Cabeçalho dos dias (1 a TOTAL_DIAS)
    col_index = 2  # Coluna B
    for dia in range(1, TOTAL_DIAS + 1):
        # Mesclar 3 colunas para o número do dia
        start_col = get_column_letter(col_index)
        end_col = get_column_letter(col_index + 2)
        ws.merge_cells(f'{start_col}{linha_cabecalho}:{end_col}{linha_cabecalho}')
        
        dia_cell = ws[f'{start_col}{linha_cabecalho}']
        dia_cell.value = dia
        dia_cell.font = Font(bold=True, size=10)
        dia_cell.alignment = Alignment(horizontal='center', vertical='center')
        dia_cell.border = thin_border
        
        # Subcolunas M, T, N na linha seguinte
        for i, turno in enumerate(['M', 'T', 'N']):
            turno_col = get_column_letter(col_index + i)
            turno_cell = ws[f'{turno_col}{linha_cabecalho+1}']
            turno_cell.value = turno
            turno_cell.font = Font(bold=True, size=9)
            turno_cell.alignment = Alignment(horizontal='center', vertical='center')
            turno_cell.border = thin_border
        
        col_index += 3
    
    # Linhas dos equipamentos
    linha_atual = linha_cabecalho + 2
    for equipamento in equipamentos:
        ws[f'A{linha_atual}'].value = equipamento
        ws[f'A{linha_atual}'].font = Font(size=9)
        ws[f'A{linha_atual}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'A{linha_atual}'].border = thin_border
        
        # Adicionar células vazias para cada dia (M, T, N)
        for col in range(2, TOTAL_COLUMNS):
            cell = ws.cell(row=linha_atual, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        linha_atual += 1
    
    # Linha "VISTO DO PROFISSIONAL"
    ws[f'A{linha_atual}'].value = "VISTO DO PROFISSIONAL"
    ws[f'A{linha_atual}'].font = Font(bold=True, size=9)
    ws[f'A{linha_atual}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'A{linha_atual}'].border = thin_border
    
    # Adicionar células vazias para a linha de visto
    for col in range(2, TOTAL_COLUMNS):
        cell = ws.cell(row=linha_atual, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 30  # Coluna MATERIAL mais larga
    
    # Colunas dos dias (M, T, N) mais estreitas
    for col in range(2, TOTAL_COLUMNS):
        ws.column_dimensions[get_column_letter(col)].width = 3
    
    # Ajustar altura das linhas
    ws.row_dimensions[1].height = 25  # Título
    ws.row_dimensions[2].height = 20  # Mês/Ano
    ws.row_dimensions[linha_cabecalho].height = 18  # Números dos dias
    ws.row_dimensions[linha_cabecalho+1].height = 18  # M, T, N
    
    for linha in range(linha_cabecalho+2, linha_atual+1):
        ws.row_dimensions[linha].height = 18
    
    # Configurar orientação da página como paisagem
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    
    # Salvar arquivo
    wb.save('CHECK_LIST_EQUIPAMENTOS_SALA_01.xlsx')
    print("Planilha 'CHECK_LIST_EQUIPAMENTOS_SALA_01.xlsx' criada com sucesso!")

if __name__ == "__main__":
    criar_checklist()
