#!/usr/bin/env python3
"""
Script para verificar a estrutura da planilha Excel gerada
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def verificar_checklist():
    wb = load_workbook('CHECK_LIST_EQUIPAMENTOS_SALA_01.xlsx')
    ws = wb.active
    
    print("=== Verificação da Planilha ===\n")
    
    # Verificar título
    titulo = ws['A1'].value
    print(f"✓ Título: {titulo}")
    
    # Verificar campo mês/ano
    mes_ano = ws['A2'].value
    print(f"✓ Mês/Ano: {mes_ano}")
    
    # Verificar cabeçalho MATERIAL
    material = ws['A3'].value
    print(f"✓ Cabeçalho: {material}")
    
    # Verificar equipamentos
    equipamentos_esperados = [
        "ELETROCAUTÉRIO",
        "MESA CIRÚRGICA TOMADA",
        "CARRO ANEST./CAPNÓGRAFO",
        "MONITOR COMPLETO",
        "OX PERFURO CORTANTE",
        "OX E LÂMINAS LARINGO",
        "ESTOJO AD E INF",
        "MÓDULO P.A.I",
        "PAREDE GASES FUNCIONANDO",
        "CIRCUITO ANESTESIA /KTS",
        "02 SUPORTE SORO",
        "03 MESAS + 01 MAYO",
        "REANIMADOR AD E INF.",
        "02 LIXEIRAS",
        "HAMPER"
    ]
    
    print(f"\n✓ Equipamentos listados:")
    linha = 5
    for i, equip in enumerate(equipamentos_esperados):
        valor_celula = ws[f'A{linha + i}'].value
        if valor_celula == equip:
            print(f"  - {equip}")
        else:
            print(f"  ✗ Esperado '{equip}', encontrado '{valor_celula}'")
    
    # Verificar linha VISTO DO PROFISSIONAL
    linha_visto = 5 + len(equipamentos_esperados)
    visto = ws[f'A{linha_visto}'].value
    print(f"\n✓ Linha final: {visto}")
    
    # Verificar dias (1 a 31) na linha 3
    print(f"\n✓ Verificando dias 1 a 31:")
    col_index = 2
    dias_encontrados = []
    for dia in range(1, 32):
        col_letter = get_column_letter(col_index)
        valor_dia = ws[f'{col_letter}3'].value
        if valor_dia == dia:
            dias_encontrados.append(dia)
        col_index += 3
    print(f"  - Dias encontrados: {len(dias_encontrados)} de 31")
    
    # Verificar turnos M, T, N na linha 4
    print(f"\n✓ Verificando turnos M, T, N:")
    col_index = 2
    turnos_ok = 0
    for dia in range(1, 32):
        m_col = get_column_letter(col_index)
        t_col = get_column_letter(col_index + 1)
        n_col = get_column_letter(col_index + 2)
        
        if (ws[f'{m_col}4'].value == 'M' and 
            ws[f'{t_col}4'].value == 'T' and 
            ws[f'{n_col}4'].value == 'N'):
            turnos_ok += 1
        
        col_index += 3
    print(f"  - Turnos corretos: {turnos_ok} de 31 dias")
    
    # Verificar orientação da página
    print(f"\n✓ Orientação da página: {'PAISAGEM' if ws.page_setup.orientation == 'landscape' else 'RETRATO'}")
    
    print("\n=== Verificação Concluída ===")

if __name__ == "__main__":
    verificar_checklist()
